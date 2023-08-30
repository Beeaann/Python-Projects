import pathlib
from pathlib import Path
import shutil
import pandas as pd
from pandas.io.formats import excel
import numpy as np
import os 
from win32com.client import constants
import win32com.client as win32
from datetime import datetime
import tkinter as tk
from tkinter import messagebox, ttk
from tkinter import *
from openpyxl import load_workbook, Workbook
import sys
import customtkinter as ctk


# Accessed Libraries
    # import pathlib \ for Path manipulation
    # from pathlib import Path
    # import shutil \ to move file to output_folder 
    # import pandas as pd \ for excel file manipulation
    # from pandas.io.formats import excel \ for excel file manipulation
    # import numpy as np \ used as a dependency for Pandas
    # import os \ to get current working directory and filename manipulation
    # from win32com.client import constants \ for excel file manipulation, pivot tables
    # import win32com.client as win32 \ for excel file manipulation, pivot tables
    # from datetime import datetime \ to add date and time to console log
    # import tkinter as tk \ dependency needed for pop-up boxes
    # from tkinter import messagebox, ttk \ use of pop-up boxes
    # from tkinter import * \ imports all necessary dependencies for faster running GUI components
    # from openpyxl import load_workbook, Workbook \ for precise (Pythonic) excel file manipulation
    # import sys \ for error handling and console log
    # import customtkinter as ctk \ for main GUI 


# Psuedo Code -
    # 1. Prompt user to select an xlsx file
        # First sheet
        # 1. Open xlsx file and look for "Sheet1"
        # 2. Copy "Sheet1" and rename copy to "Filtered" then rename "Sheet1" to "original"
        # 3. Remove rows with "ACH, Cash, Payment Vision, wire, direct to client" from "Transaction method" column, 
            # but keep "Cashiers check"
        # 4. Re-order column A to ascending order
        # 5. Split client name to main name using "Client Names for PM report.xlsx", example: "calvary-sc" to "Calvary"
            # Compare cell value in "Filtered" sheet against column A's contents in client name report file
                # If cell matches, replace value in cell with value from client name report
                # If not, continue
            # (Due to pandas formatting, operations are in place to fix datetime conversion of dates and accounting formats)
    # 2. Create pivot table named "Payment Count Review"
        # Add "Claim Number" column from "Original" sheet
        # Then add second column called "Count of claim number", displaying amount of occurrences within
        # each cell of Claim numbers. 
        # Add grand total of claim numbers counted to bottom row
        # Highlight any cells that are greater than 1
    # 3. Create another pivot table called "Balance"
        # Import "Client Name" column from "Filtered" sheet
        # Add second column "Sum of amount" == Summed up "Amount" column in "Filtered" sheet
        # Add Grand total row to "Client Name" and "Sum of amount"
        # Add "Pivot" column containing duplicate information as in "Sum of amount"
        # Add "Checks" column containing nothing for manual check information insertion
        # Add "Diff" column containing formula: "Checks" - "Diff"
            # Rename "Diff" header back to "Diff" from win32com formula add on name
        # Add duplicate client names two rows below grand total row
        # Add two sections as follows: |"Sticky Note Total" | =SUM(B12:B25)|
        #                              |"Diff"              |  =B26-B11    |
            # (Note: Equations are placed based on last row after grand total row)
    # 4. Save and close file, renaming it to date within "Entered Date" column in "Filtered" sheet.
    # 5. Create output_folder and move file to it.
        
        
        
# Set style of headers for all excel manipulation
excel.ExcelFormatter.header_style = None

# Function to create "Filtered" table (see pseudo code)
def addFilteredTable(fileName):
    
    # Enclose function in try\except for continuity purposes
    try:
        
        # Get current working directory
        cwd = os.getcwd()
        
        # Create new file name
        XLFile = os.path.join(cwd, fileName)
        
        #launch excel application
        xlApp = win32.gencache.EnsureDispatch("Excel.Application")
        xlApp.Visible = False

        #create workbook : wb and ws
        #reference workbooks
        wb = xlApp.Workbooks.Open(XLFile)

        #reference worksheets
        ws_Original = wb.Worksheets("Filtered")
        ws_Balance = wb.Worksheets("Balance") 

        #create pivot table cache connection 
        pt_cache = wb.PivotCaches().Create(1,ws_Original.Range("B:S").CurrentRegion)

        #create pivot table designer/editor
        pt = pt_cache.CreatePivotTable(ws_Balance.Range("A3"), "myreport_summary")
       
        #row and column grandtotals
        pt.ColumnGrand = True
        pt.RowGrand = True

        #change report layout
        pt.RowAxisLayout(1)   #RowAxisLayout(1) for tabular form

        #change pivot table style
        #Select from Design tab, try out Medium9 or Medium3
        pt.TableStyle2 = "pivotStyleMedium9"   
        
        #add cols
        field_cols = {} 
        field_cols["Client Name"] = pt.PivotFields("Client Name")
        
        #insert col field to pivot table design
        field_cols["Client Name"].Orientation = 1   # 2 for col orientation

        
        #Add values to "Client Name" and "Sum of amount" columns 
        # from "Client Name" and "Amount" columns in "Original" sheet.
        field_values = {} 
        field_values["Client Name"] = pt.PivotFields("Client Name")
        field_values["Amount"] = pt.PivotFields("Amount")
        field_values['Amount'].Orientation = 4
        field_values['Amount'].Function = -4157 # value reference
        field_values['Amount'].NumberFormat = "#,##0.00"
        
        # Get the pivot table's source data range
        source_range = pt.TableRange1

        # Find the column position of the "Amount" column
        amount_column = source_range.Find("Amount")
        amount_column_number = amount_column.Column

        # Calculate the target row for insertion (start at third row)
        target_row = 3

        # Insert a new column after the "Amount" column, starting from the target row
        new_column_number = amount_column_number + 1
        new_column_range = ws_Balance.Columns(new_column_number)
        new_column_range.Insert(Shift=-4161)  # Shift cells down

        # Set the header for the new column at the target row
        new_column_header = ws_Balance.Cells(target_row, new_column_number)
        new_column_header.Value = "Pivot"

        # Insert another new column after the "Checks" column, starting from the target row
        new_column_number += 1
        new_column_range = ws_Balance.Columns(new_column_number)
        new_column_range.Insert(Shift=-4161)  # Shift cells down

        # Set the header for the new column at the target row
        new_column_header = ws_Balance.Cells(target_row, new_column_number)
        new_column_header.Value = "Checks"

        # Add "Diff" column
        new_column_number += 1
        new_column_range = ws_Balance.Columns(new_column_number)
        new_column_range.Insert(Shift=-4161)  # Shift cells down

        # Set the header for the new column at the target row
        new_column_header = ws_Balance.Cells(target_row, new_column_number)
        new_column_header.Value = "Diff"
        
        # Refresh sheet
        pt.PivotCache().Refresh()
        
        # Get last row location
        last_row = ws_Balance.Cells(ws_Balance.Rows.Count, amount_column_number).End(-4162).Row
        
        # Get the range for "Sum of Amount" column
        sum_of_amount_column_range = ws_Balance.Range(ws_Balance.Cells(target_row + 1, amount_column_number),
                                                    ws_Balance.Cells(last_row, amount_column_number))

        # Get the range for the "Pivot" column
        pivot_column_range = ws_Balance.Range(ws_Balance.Cells(target_row + 1, new_column_number - 2),
                                            ws_Balance.Cells(last_row - 1, new_column_number - 2))

        # Set the values from "Sum of Amount" column to "Pivot" column
        pivot_column_range.Value = sum_of_amount_column_range.Value

        # Create formula for "Diff" column
        diff_formula_range = ws_Balance.Range(
        ws_Balance.Cells(target_row, new_column_number),
        ws_Balance.Cells(last_row - 1, new_column_number))
        
        # Apply formula to each cell in "Diff" column
        for i in range(last_row - target_row + 1):
            check_cell = ws_Balance.Cells(target_row + i, new_column_number - 1)
            pivot_cell = ws_Balance.Cells(target_row + i, new_column_number - 2)
            diff_formula_range.Cells(i + 1).Formula = f'=IFERROR({check_cell.Address}-{pivot_cell.Address},"Diff")'
        
        # Empty last cell in column (win32 com adds empty cell to match grand total, but not necessary)
        last_diff_cell = ws_Balance.Cells(last_row, new_column_number)
        last_diff_cell.ClearContents() 
       
        # Find last row in column 'A'
        last_row = ws_Balance.Range('A' + str(ws_Balance.Cells(ws_Balance.Rows.Count, "A").End(-4162).Row)).Row  
        first_data_row = 4  # Assuming header is in the first row, so data begins at the second row

        # Calculate destination column
        destination_first_row = last_row + 3  # Two rows below the last row plus one row for the new header

        # Copy data without header and last row
        range_to_copy = ws_Balance.Range(f"A{first_data_row}:A{last_row - 1}")
        range_to_copy.Copy()

        # Paste data two rows below the original client name column
        destination_range_top_left_cell = ws_Balance.Range(f"A{destination_first_row}")
        destination_range_top_left_cell.PasteSpecial()
         
        # Apply Accounting number format to cell B26 and write title next to cell called 'Sticky Note Totals'
        ws_Balance.Range("B26").NumberFormat = "_($* #,##0.00_);_($* (#,##0.00);_($* \"-\"??_);_(@_)"
        ws_Balance.Range("A26").Value = 'Sticky Note Totals'
        
        # Apply bold format to cell - C3
        boldRange1 = ws_Balance.Range("C3")
        boldRange1.Font.Bold = True
        boldRange1.HorizontalAlignment = win32.constants.xlCenter
        
        # Apply bold format to cell - D3
        boldRange2 = ws_Balance.Range("D3")
        boldRange2.Font.Bold = True
        boldRange2.HorizontalAlignment = win32.constants.xlCenter
        
        # Apply bold format to cell - E3
        boldRange3 = ws_Balance.Range("E3")
        boldRange3.Font.Bold = True
        boldRange3.HorizontalAlignment = win32.constants.xlCenter
        
        # Apply bold format to cell - A26
        boldRange5 = ws_Balance.Range("A26")
        boldRange5.Font.Bold = True
        boldRange5.HorizontalAlignment = win32.constants.xlCenter
        
        # Apply bold format to cell - A28
        boldRange5 = ws_Balance.Range("A28")
        boldRange5.Font.Bold = True
        boldRange5.HorizontalAlignment = win32.constants.xlCenter
        
        # Apply bold format to cell - B26
        boldRange4 = ws_Balance.Range("B26")
        boldRange4.Font.Bold = True
        boldRange4.HorizontalAlignment = win32.constants.xlCenter
        
        # Apply bold format to cell - B28
        boldRange5 = ws_Balance.Range("B28")
        boldRange5.Font.Bold = True
        boldRange5.HorizontalAlignment = win32.constants.xlCenter
        
        # Identify B26 range for border application
        range1 = ws_Balance.Range("B26")
        borders = range1.Borders

        # Apply top border to B26
        borders(win32.constants.xlEdgeTop).LineStyle = win32.constants.xlContinuous
        borders(win32.constants.xlEdgeTop).Weight = win32.constants.xlThin

        # Apply bottom border to B26
        borders(win32.constants.xlEdgeBottom).LineStyle = win32.constants.xlContinuous
        borders(win32.constants.xlEdgeBottom).Weight = win32.constants.xlThin
        
        # Identify B28 range for border application
        range2 = ws_Balance.Range("B28")
        borders2 = range2.Borders

        # Apply top border to B28
        borders2(win32.constants.xlEdgeTop).LineStyle = win32.constants.xlContinuous
        borders2(win32.constants.xlEdgeTop).Weight = win32.constants.xlThin

        # Apply bottom border to B28
        borders2(win32.constants.xlEdgeBottom).LineStyle = win32.constants.xlContinuous
        borders2(win32.constants.xlEdgeBottom).Weight = win32.constants.xlThin
        
        # Add title in A28 called "Diff."
        ws_Balance.Range("A28").Value = 'Diff.'
        
        
        # for loop to start from the last cell in column A and move upwards
        last_row = ws_Balance.UsedRange.Rows.Count
        total_cell = None
        for row in range(last_row, 0, -1):
            cell = ws_Balance.Cells(row, 1)  # 1 is for column A
            # If cell found has "Grand total" then save the cell next to it(Grand total value)
            if cell.Value == "Grand Total":
                total_cell = ws_Balance.Cells(row, 2)  # 2 is for column B
                break
            
        # Error handling for if not found
        if total_cell is None:
            print("No 'Grand total' cell found in column A.")
        else:
            # Add formula to B28 in case Grand total value is set
            formula_cell = ws_Balance.Range("B28")
            formula_cell.Formula = f"=B26-B{total_cell.Row}"
            ws_Balance.Range("B26").Formula = f"=SUM(B{total_cell.Row + 1}:B25)"
        
        # *** FILTERED SHEET ***
        #       |   |   |
        #       V   V   V
        
        # Creation of a 2x2 total posted values and checks in filtered sheet
        
        # Get the last row in the spreadsheet
        last_row = ws_Original.UsedRange.Rows.Count

        # Calculate the number of rows in the spreadsheet (minus the header)
        posted_value = last_row - 1  # adjust accordingly if you have more than one header row
        checks_value = ""
        
        # Insert 'Posted' and 'Checks' two rows after the last row under the columns "R" and "S"
        ws_Original.Cells(last_row + 3, 18).Value = 'Posted'
        ws_Original.Cells(last_row + 4, 18).Value = 'Checks'
       
        
        # Insert the posted_value next to 'Posted'
        ws_Original.Cells(last_row + 3, 19).Value = posted_value
        ws_Original.Cells(last_row + 3, 19).NumberFormat = '0'
        
        # Apply number format to empty cell 
        ws_Original.Cells(last_row + 4, 19).Value = checks_value
        ws_Original.Cells(last_row + 4, 19).NumberFormat = '0'
        
        # Add a formula to calculate the difference between 'Posted' and 'Checks'
        ws_Original.Cells(last_row + 5, 19).Formula = f"={ws_Original.Cells(last_row + 3, 19).GetAddress(False, False)}-{ws_Original.Cells(last_row + 4, 19).GetAddress(False, False)}"
        ws_Original.Cells(last_row + 5, 19).NumberFormat = '0'
        
        # Add total cell below second "Column"
        total_cell = ws_Original.Cells(last_row + 5, 19)

        # Add border around the cell
        border = total_cell.Borders
        for side in [constants.xlEdgeTop, constants.xlEdgeBottom]:
            border_item = border(side)
            border_item.LineStyle = constants.xlContinuous
            border_item.Weight = constants.xlThin
            
        # Save and close file
        wb.Save()
        wb.Close(True)
    
    # Raise exception displaying type error, line occurence, and name of error
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(f"Exception type - {exc_type},\n, File Name - {fname}\n, Line Number - {exc_tb.tb_lineno}\n")
    
    # Finally, quit program so excel is not stuck running in background 
    finally:
        xlApp.Quit() 
        
# Function to add Payment Count Review(PCR) Pivot table (See pseudocode)
def addPCRTable(fileName):


    # Enclose function in try\except for continuity purposes
    try:
        
        #Get current working directory
        cwd = os.getcwd()
        
        # Create new filename
        XLFile = os.path.join(cwd, fileName)
        
        #launch excel application
        xlApp = win32.gencache.EnsureDispatch("Excel.Application")
        xlApp.Visible = False

        #create workbook : wb and ws
        #reference workbooks
        wb = xlApp.Workbooks.Open(XLFile)

        #reference worksheets
        ws_Original = wb.Worksheets("Original")
        ws_Pivot = wb.Worksheets("Payment Count Review") 
        
      
        #create pivot table cache connection 
        pt_cache = wb.PivotCaches().Create(1,ws_Original.Range("E:E").CurrentRegion)

        #create pivot table designer/editor
        pt = pt_cache.CreatePivotTable(ws_Pivot.Range("A3"), "myreport_summary")

        #row and column grandtotals
        pt.ColumnGrand = True
        pt.RowGrand = True

        #change report layout
        pt.RowAxisLayout(1)   #RowAxisLayout(1) for tabular form

        #change pivot table style
        #Select from Design tab, try out Medium9 or Medium3
        pt.TableStyle2 = "pivotStyleMedium9"   
        
        #add cols
        field_cols = {} 
        field_cols["Claim Number"] = pt.PivotFields("Claim Number")
        
        #insert col field to pivot table design
        field_cols["Claim Number"].Orientation = 1   # 2 for col orientation

        #add value
        field_values = {} 
        field_values["Claim Number"] = pt.PivotFields("Claim Number")
        
        #insert data/value to pivot table design
        field_values["Claim Number"].Orientation = 4   # 4 for data/value
        field_values["Claim Number"].Function = -4112  # -4112 for xlCount
        field_values["Claim Number"].NumberFormat = "#,##0" # "#,##0" for number format
        
        # Refresh sheet
        pt.PivotCache().Refresh()

        # Identify last row
        lastRow = ws_Pivot.Cells(ws_Pivot.Rows.Count, "B").End(-4162).Row

        # Set the range you want to apply conditional formatting to
        cond_range = ws_Pivot.Range(f"B4:B{lastRow-1}")

        # Add the conditional formatting
        cond_format = cond_range.FormatConditions.Add(Type=1, Operator=4, Formula1="=1")  # Type=1 is xlCellValue, Operator=4 is xlGreater

        # Set color of highlighted cells
        light_blue_color = 52 + 173*256 + 239*65536
        cond_format.Interior.Color = light_blue_color
        
        # Save and close workbook
        wb.Save()
        wb.Close(True)
    
    # Raise exception displaying type error, line occurence, and name of error
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(f"Exception type - {exc_type},\n, File Name - {fname}\n, Line Number - {exc_tb.tb_lineno}\n")
    
    # Finally, quit program so excel is not stuck running in background 
    finally:
        xlApp.Quit() 

# Function to split client names, renaming "Client Name" cell values with values in
# "Client Names for PM report.xlsx"
def splitClientNames(filteredSheet, rowAmount):
    
    # Read values from 'Client names for PM report.xlsx'
    clientNameFileDF = pd.read_excel('Client names for PM report.xlsx')

    # Zip values as dictionary for comparison 
    clientNameDict = dict(zip(clientNameFileDF['Client Id'], clientNameFileDF['Client']))

    # Actual comparison loop
    i = 0
    while i < rowAmount:
        
        for key, value in clientNameDict.items():
            
            if filteredSheet['Client ID'].iloc[i] == key: # Compare cell in "Client ID" 
                filteredSheet['Client Name'].iloc[i] = value # against cell in "Client Name"
                print(f"Removing {key} for {value}")
        i += 1 

# Function to remove specific transaction methods and their rows from "Transaction Method" column
def removeTransactionMethods(passedFile):
    
    # List of things that need to be removed from "Transaction Method" column
    removeList = ["ACH", "Cash", "Payment", "Wire", "Direct", "Out", "Reverse", "Offset"]
    
    # Load workbook
    book = load_workbook(passedFile)
    
    # Identify "Original" sheet
    sheet1 = book.worksheets[0]
    sheet1.title = "Original"
    
    # Make copy of "Original" sheet
    book.copy_worksheet(sheet1)
    
    # Identify "Filtered" sheet
    sheet2 = book.worksheets[1]
    sheet2.title = "Filtered"
    
    # Create new sheet called "Payment Count Review"
    piv1 = book.create_sheet("Payment Count Review")
    piv1.title = "Payment Count Review"
    
    # Create new sheet called "Balance"
    piv2 = book.create_sheet("Balance")
    piv2.title = "Balance"
    
    # Create new name for file (for honed in manipulation)
    alteredBook = f"{Path(passedFile).stem}_alt.xlsx"
    
    # While loop to remove transaction methods
    i = 1
    while i <= sheet2.max_row:
        
        # If cell is empty, continue
        if sheet2.cell(row=i, column=12).value == None:
            continue
        
        # elif cell value == "Cashiers Check", skip it and keep going (Keep cashiers check)
        elif sheet2.cell(row=i, column=12).value == "Cashiers Check":
            i += 1
            continue
        
        # elif cell value is something in removeList, remove row with that item
        elif sheet2.cell(row=i, column=12).value.startswith(tuple(removeList)):
            sheet2.delete_rows(i, 1)
        
        # Add 1 to counter
        else: 
            i += 1
    
    # Save file
    book.save(alteredBook)

# Function to do bulk of file manipulation (PM Process)
def excelRunner(currentFile):
    
    # Initialize output_folder
    pathname = os.path.dirname(sys.argv[0])  
    output_folder_name = "output"
    output_dir = pathlib.Path(pathname) / output_folder_name
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Open console log for writing
    sys.stdout = open(f"{output_dir}\\ConsolePM.txt", 'w')
    
    # Display all pandas output in 3 decimal place format
    pd.options.display.float_format = '{:.3f}'.format
    
    # Call to removeTransactionMethods with passing of currentFile which is file name
    removeTransactionMethods(currentFile)
    
    # Create new name for file (for honed in manipulation)
    alteredBook = f"{Path(currentFile).stem}_alt.xlsx"
    
    print("Getting Dates")
    
    # Call on getDate function, with passing of currentFile which is file name
    # Get date from "Entered date column"
    getDate(currentFile)
    
    # Load new name file
    df = pd.ExcelFile(alteredBook)

    # Create dataframe from new name file
    realDF = pd.read_excel(alteredBook, sheet_name=None)
    
    # Identify all four sheets
    sheet1_dataFrame = pd.read_excel(df, 'Original')
    sheet2_DataFrame = pd.read_excel(df, 'Filtered')
    PCR_DataFrame = pd.read_excel(df, 'Payment Count Review')
    balance_DataFrame = pd.read_excel(df, 'Balance')
 
    
    # Get total length of columns and rows
    columnAmount= len(sheet2_DataFrame.columns)
    rowAmount = len(sheet2_DataFrame)
    
    # Display total length of columns and rows
    pd.set_option('display.max_columns', columnAmount)
    pd.set_option('display.max_rows', rowAmount)
    

    # Function call splitClientNames with passing of "Filtered" sheet dataframe and total amount of rows
    splitClientNames(sheet2_DataFrame, rowAmount)
    
    # Create sorted dataframe of "Filtered" sheet, sorting by "Client ID" column in ascending order
    sortedDF = sheet2_DataFrame.sort_values('Client ID', ascending=True)

    
    # *** FILTERED DATAFRAME (SORTED) ***
    #       |   |   |
    #       V   V   V
    
    # get all header names as list  
    headerNames = list(sortedDF.columns.values)

    # Retrieve all column headers with dates
    headersWithDates = []
    for header in headerNames:
  
        if header.endswith("Date"):
            headersWithDates.append(header)

    # for loop to convert each column with dates back to normal date format
        # Note: Pandas converts any date or time into datetime format upon reading file
        # for this program's purposes, we need to preserve date format
    for item in headersWithDates:
        
        # if cell data type is datetime object 
        if sortedDF[item].dtype == 'datetime64[ns]':
            
            # Convert any datetime values to string type
            sheet1_dataFrame[item] = sheet1_dataFrame[item].astype(str)
            sortedDF[item] = sortedDF[item].astype(str)
            
            # Drop time stamps from "Original" sheet and change formatting of date back
            # to original format
            sheet1_dataFrame[item] = pd.to_datetime(sheet1_dataFrame[item])
            sheet1_dataFrame[item] = sheet1_dataFrame[item].dt.strftime('%m/%d/%Y')
            
            # Drop time stamps from "Filtered"(sortedDF) sheet and change formatting of date back
            # to original format    
            sortedDF[item] = pd.to_datetime(sortedDF[item])
            sortedDF[item] = sortedDF[item].dt.strftime('%m/%d/%Y')
            
            print(f"Column:{item} is a datetime data type")
            
        else:
            print(f"Column{item} is something else!")
        
        # Convert 'Claim Number', 'Network Number' columns to string format for proper display 
        # within excel file
        sheet1_dataFrame[['Claim Number', 'Network Number']] = sheet1_dataFrame[['Claim Number', 'Network Number']].astype(str)
        sortedDF[['Claim Number', 'Network Number']] = sortedDF[['Claim Number', 'Network Number']].astype(str)
    
    # Create list of all columns in "Original" and "Filtered" sheets that contain money values
    # for re-conversion to accounting format
    moneyList = ['Amount', 'Principal', 'Interest', 'Fees', 'Costs', 'Other', 'Client Proceeds', 'Commission', 'Direct Payment', 'Contingent Suit Fee', 'Invoice Pmt', 'Current Balance', 'Open Balance']

    # Create sheet dictionary for saving and formatting purposes ('Sheet Name': sheetVar)
    sheetDict = {'Original':sheet1_dataFrame, 'Filtered':sortedDF, 'Payment Count Review':PCR_DataFrame, 'Balance':balance_DataFrame}  
    with pd.ExcelWriter(alteredBook, engine='xlsxwriter') as writer:
    
        # loop through `dict` of dataframes
        for sheetname, df in sheetDict.items():  
            
            # send df to writer
            df.to_excel(writer, sheet_name=sheetname, index=False)  
            worksheet = writer.sheets[sheetname]  # pull worksheet object
            
            # loop through all columns to resize columns to length of longest cellValue
            for idx, col in enumerate(df):  
                series = df[col]
                max_len = max((
                    series.astype(str).map(len).max(),  # len of largest item
                    len(str(series.name))  # len of column name/header
                )) + 1  # adding a little extra space
                worksheet.set_column (idx, idx, max_len)  # set column width

                # If this column is in your moneyList, apply the accounting format.
                if col in moneyList:
                    money_format = writer.book.add_format({'num_format': '#,##0.00'})
                    worksheet.set_column(idx, idx, None, money_format)
  
    print("\n\nsheet names - ", realDF.keys())   

    # Function calls to add "Filtered" and "Payment Count Review" pivot tables, passing fileName with them 
    addFilteredTable(alteredBook)
    addPCRTable(alteredBook)
    
    # try/except block to create new final file name 
    try:
        newName = f"{year} {month} {day} Payments.xlsx" #create new name
        newPath = os.path.join(output_dir, newName) # conjoin new name with output_folder path
        shutil.move(alteredBook, newPath) # move new file to output_folder
    
    # show error if any
    except Exception as error:
        print("Error", error)
    
    # Close console log object    
    print("Finished")
    sys.stdout.close()

# Function to get exact date from "Entered Date" column in "Original" sheet 
def getDate(df):
    
    # Global variables and read "Sheet1"
    global month, day, year
    df = pd.read_excel(df, sheet_name='Sheet1')
    
    # Identify "Entered Date" column and get second row value, then convert it to string
    data = df['Entered Date'][2]
    data = str(data)
    
    # Create year month and day values based on string cell 
    year = data[0:4]
    month = data[5:7]
    day= data[8:10]

# "Main" function that calls on "excelRunner" function, which runs rest of program  
def PM_Run():
    
    # Get current working directory 
    cwd = os.getcwd()

    # try/except block to prompt user for file then run "excelRunner" function,
    # passing name of user-selected file
    try:
        file = ctk.filedialog.askopenfilename(title="Select a file", initialdir=cwd)
        excelRunner(file)
    
    # Display error if any
    except Exception as error:
        print(error)
    
        
        
#Calls main function in program, loop to run window infinitely.
if __name__ == "__main__":
    PM_Run()     