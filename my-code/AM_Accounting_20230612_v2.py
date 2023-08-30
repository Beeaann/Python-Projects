import pathlib #used to get path
import shutil #used to move file
import pandas as pd #used to manipulate Excel files
from pandas.io.formats import excel #used for money format 
from pathlib import Path #used for getting path
import os #used for CWD
import win32com.client as win32 #used for pivot table 
from win32com.client import constants #used for pivot table
from datetime import datetime #used for date
from openpyxl import load_workbook #used for
import customtkinter as ctk #used for GUI
import sys 
from PIL import Image #used for GUI image
from PM_Accounting import PM_Run #open another py file for PM Process
from time import sleep #Pauses program

# Psuedo Code -
    # 1. Read each .xlsx file in directory.
    # 2. Return the dataframe.
        # -Allow user to pick what needs to be removed from "Transaction Method" column
            # - Maybe using notepad, buttons, or entry box
    # 3. Remove rows with 
    # 3. Re-order contents in column A to alphabetical order
    # 4. Split client name to main name
    # 5. Create Grand Total box on original sheet
    # 6. Create Pivote Table
  
def gui():
    global root, textbox, progress, option, button1
    ctk.set_appearance_mode("Dark")
    ctk.set_default_color_theme("dark-blue")
    root = ctk.CTk()
    root.title("AM Account")
    root.resizable(False, False)
    window_width = 550
    window_height = 550

    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
        # Coordinates of the upper left corner of the window to make the window appear in the center
    x_cordinate = int((screen_width/2) - (window_width/2))
    y_cordinate = int((screen_height/2) - (window_height/2))
    root.geometry("{}x{}+{}+{}".format(window_width, window_height, x_cordinate, y_cordinate))
    
    mainFrame = ctk.CTkFrame(root)
        
    datafile2 = "logo.png"
    if not hasattr(sys, "frozen"):
        datafile2 = os.path.join(os.path.dirname(__file__), datafile2)
    else:
        datafile2 = os.path.join(sys.prefix, datafile2)
    image= ctk.CTkImage(Image.open(datafile2), size=(150,115))
    img_label = ctk.CTkLabel(root, image=image, text="")
    img_label.pack()
    
    mainFrame.pack(side="top", fill="both", expand=True, padx=25, pady=25)
    
    progress = ctk.CTkProgressBar(mainFrame, mode="indeterminate", width=425)
    
    textbox = ctk.CTkTextbox(mainFrame, width=450, height=300)
    textbox.pack(pady=5)
    textbox.configure(state="disabled")
    filetype = ctk.StringVar(value="AM Accounting")
    
    button1 = ctk.CTkButton(mainFrame, command=main, text="Start Process", width=140)
    button1.pack(pady= 5, side="bottom") 
    option = ctk.CTkOptionMenu(mainFrame, values=['AM Accounting','PM Accounting'], command=optionSelection, variable=filetype)
    option.pack(side='bottom', pady=5)
 
excel.ExcelFormatter.header_style = None     
def getDate(df):
    global month, day
    df = pd.read_excel(df, sheet_name='Sheet1')
    data = str(df['Entered Date'][2])
    try:
        date = datetime.strptime(data, '%Y-%m-%d %H:%M:%S')
        month = date.strftime('%m')
        day = date.day
    except Exception as error:
        print("Trying Different Date Format")
        month = data[0:2]
        day= data[6:8]
     
    
    print(data)
    print("month: ", month)
    print("day: ", day)
    date = f"{month}{day}"
    return "Filtered"

    #used to output to text file selection
def optionSelection(choice):
    textbox.configure(state="normal")
    textbox.insert(ctk.END, f"Selected: {choice}\n")
    textbox.configure(state="disabled")    
    
    #used to add pivot table
def addFilteredTable(wb):
    textbox.configure(state="normal")
    cwd = os.getcwd()
         
    #print("cwd from addpivot table - ", cwd)
    XLFile = os.path.join(cwd, wb)
    
    #launch excel application
    xlApp = win32.gencache.EnsureDispatch("Excel.Application")
    xlApp.Visible = False

    #create workbook : wb and ws
    #reference workbooks
    wb = xlApp.Workbooks.Open(XLFile)
    sleep(1)
    
    #reference worksheets
    ws_first = wb.Worksheets('Original')
    ws_Original = wb.Worksheets(date)
    ws_Balance = wb.Worksheets("Balance") 
    ws_Balance.Name = "Recon"
    #create pivot table cache connection 
    pt_cache = wb.PivotCaches().Create(1,ws_Original.Range("B:S").CurrentRegion)

    #create pivot table designer/editor
    pt = pt_cache.CreatePivotTable(ws_Balance.Range("A3"), "myreport_summary")
    
    #row and column grandtotals
    pt.ColumnGrand = True
    pt.RowGrand = True

    #change report layout
    pt.RowAxisLayout(0)   #RowAxisLayout(1) for tabular form

    #change pivot table style
    #Select from Design tab, try out Medium9 or Medium3
    pt.TableStyle2 = "pivotStyleMedium9"   
    
    #add cols
    field_cols = {} 
    field_cols["Transaction Type"] = pt.PivotFields("Transaction Type")
    
    #insert col field to pivot table design
    field_cols["Transaction Type"].Orientation = 2   # 2 for col orientation
    
    field_rows = {}
    field_rows["Client Name"] = pt.PivotFields("Client Name")
    field_rows["Transaction Method"] = pt.PivotFields("Transaction Method")
    field_rows["Client Name"].Orientation = 1
    field_rows["Transaction Method"].Orientation = 1
    
    #add value
    field_values = {} 
    field_values["Amount"] = pt.PivotFields("Amount")
    
    field_values['Amount'].Orientation = 4
    field_values['Amount'].Function = -4157 # value reference
    field_values['Amount'].NumberFormat = "#,##0.00"
    
    pt.PivotCache().Refresh()
       
    last_row = ws_first.Cells(ws_first.Rows.Count, "A").End(-4162).Row
    total = 0
    for i in range(2, last_row + 1):
        amount = ws_first.Cells(i, 19).Value
        if isinstance(amount, str):
            amount = int(amount)
        total += amount
        
    ws_first.Cells(last_row + 2, 19).Value = total
    ws_first.Columns(19).AutoFit()
    
    border_range = ws_first.Range(ws_first.Cells(last_row + 2, 19), ws_first.Cells(last_row + 2, 19))
    border_range.Borders(constants.xlEdgeTop).LineStyle = constants.xlContinuous
    border_range.Borders(constants.xlEdgeBottom).LineStyle = constants.xlDouble
    border_range.Borders(constants.xlEdgeBottom).Weight = constants.xlThick

    wb.Save()
    wb.Close(True)
    
    #Get list of client names from AM Excel file to modify main file client names
def splitClientNames(df, rowAmount):
    
    print("Current directory split client-", os.getcwd())
    textbox.configure(state="normal")
    clientNameFileDF = pd.read_excel('Client names for AM report.xlsx')
 
    clientNameDict = dict(zip(clientNameFileDF['Client Id'], clientNameFileDF['Client']))

    i = 0
    while i < rowAmount:
        
        for key, value in clientNameDict.items():
            
            if df['Client ID'].iloc[i] == key:
                df['Client Name'].iloc[i] = value
          
        i += 1 
    textbox.configure(state="disabled")
    return df
    
    #remove selected transaction methods along with row
def removeTransactionMethods(passedFile):
    global date
    date = getDate(passedFile)
    
    textbox.configure(state="normal")
    removeList = ["Direct", "NEEDSECONDFIELD"]
    book = load_workbook(passedFile)
    
    sheet1 = book.worksheets[0]
    sheet1.title = "Original"
    textbox.insert(ctk.END, f"'Original' Sheet Created!\n")
    book.copy_worksheet(sheet1)
    
    sheet2 = book.worksheets[1]
    sheet2.title = date
    textbox.insert(ctk.END, f"{date} Sheet Created!\n")
    
    piv2 = book.create_sheet("Balance")
    piv2.title = "Balance"
    textbox.insert(ctk.END, f"'Balance' Sheet Created!\n")
    
    alteredBook = f"{Path(passedFile).stem}_alt.xlsx"
    textbox.insert(ctk.END, f"Empty Excel File Created!\n")
   
    i = 1
    total = 0
    while i <= sheet2.max_row:
        
        if sheet2.cell(row=i, column=13).value == None:
            continue
        elif sheet2.cell(row=i, column=13).value.startswith(tuple(removeList)):
            sheet2.delete_rows(i, 1)
            textbox.insert(ctk.END, f"Deleting 'Direct to Client' on ({i}, 13)\n")
            print("Deleting Direct to Client")
        else: 
            i += 1  
    
    book.save(alteredBook)
    book.close()

#Runs process of manipulating excel file
def excelRunner(currentFile):
    textbox.configure(state="normal")
    
    pd.options.display.float_format = '{:.3f}'.format
    
    removeTransactionMethods(currentFile)
    try:
        pathname = os.path.dirname(sys.argv[0])  
        output_folder_name = "output"
        output_dir = pathlib.Path(pathname) / output_folder_name
        output_dir.mkdir(parents=True, exist_ok=True)
    except Exception as error:
        print(error)
        
    sys.stdout = open(f"{output_dir}\\Console({date}).txt", 'w')
    alteredBook = f"{Path(currentFile).stem}_alt.xlsx"
    print("AM Process")
    df = pd.ExcelFile(alteredBook)
    print("Creating Sheets")
    sheet1_dataFrame = pd.read_excel(df, 'Original')
    sheet2_DataFrame = pd.read_excel(df, date)
    balance_DataFrame = pd.read_excel(df, 'Balance')
    
    columnAmount= len(sheet2_DataFrame.columns)
    rowAmount = len(sheet2_DataFrame)
    pd.set_option('display.max_columns', columnAmount)
    pd.set_option('display.max_rows', rowAmount)
    
    splitClientNames(sheet2_DataFrame, rowAmount)
    
    sortedDF = sheet2_DataFrame.sort_values('Client ID', ascending=True)

    headerNames = list(sortedDF.columns.values)
    textbox.insert(ctk.END, f"Fixing Dates!\n")
    
    headersWithDates = []
    for header in headerNames:
        if header.endswith("Date"):   
            headersWithDates.append(header)
    print("Fixing Dates")
    for item in headersWithDates:
        if sortedDF[item].dtype == 'datetime64[ns]':
            
            sheet1_dataFrame[item] = sheet1_dataFrame[item].astype(str)
            sortedDF[item] = sortedDF[item].astype(str)
            
            sheet1_dataFrame[item] = pd.to_datetime(sheet1_dataFrame[item])
            sheet1_dataFrame[item] = sheet1_dataFrame[item].dt.strftime('%m/%d/%Y')

            sortedDF[item] = pd.to_datetime(sortedDF[item])
            sortedDF[item] = sortedDF[item].dt.strftime('%m/%d/%Y')
            
            print(f"Column:{item} is a datetime data type")
            
        else:
            print(f"Column {item} is something else!")
        
        sheet1_dataFrame[['Claim Number', 'Network Number']] = sheet1_dataFrame[['Claim Number', 'Network Number']].astype(str)
        sortedDF[['Claim Number', 'Network Number']] = sortedDF[['Claim Number', 'Network Number']].astype(str)
    
    sheetDict = {'Original':sheet1_dataFrame, date:sortedDF, "Balance":balance_DataFrame} 
    print("Exporting File") 
    moneyList = ['Amount', 'Principal', 'Interest', 'Fees', 'Costs', 'Other', 'Client Proceeds', 'Commission', 'Direct Payment', 'Contingent Suit Fee', 'Invoice Pmt', 'Current Balance', 'Open Balance']
    textbox.insert(ctk.END, f"Writing Data to Excel file\n")
    with pd.ExcelWriter(alteredBook, engine='xlsxwriter') as writer:
        print("Fixing Money Format")
        for sheetname, df in sheetDict.items():  # loop through `dict` of dataframes
            df.to_excel(writer, sheet_name=sheetname, index=False)  # send df to writer
            worksheet = writer.sheets[sheetname]  # pull worksheet object
            
            for idx, col in enumerate(df):  # loop through all columns
                series = df[col]
                max_len = max((
                    series.astype(str).map(len).max(),  # len of largest item
                    len(str(series.name))  # len of column name/header
                    )) + 1  # adding a little extra space
                worksheet.set_column (idx, idx, max_len)  # set column width  
                if col in moneyList:
                    money_format = writer.book.add_format({'num_format': '#,##0.00'})
                    worksheet.set_column(idx, idx, None, money_format)
            
    textbox.insert(ctk.END, f"Creating Pivot Table!\n")
    textbox.insert(ctk.END, f"Creating Grand Total in Amount Column\n")
    print("Creating Pivot Table")
    addFilteredTable(alteredBook)
    newName = f"{month}{day}.xlsx"
    newPath = os.path.join(output_dir, newName)
    textbox.insert(ctk.END, f"Renaming file to : {newName}\n")
    textbox.insert(ctk.END, f"Moving file to: {newPath}\n")
    shutil.move(alteredBook, newPath)
    textbox.insert(ctk.END, f"Process Complete!\n")
    print("Finished")
    sys.stdout.close()
    
def main():
    textbox.configure(state="normal")
    cwd = os.getcwd()
    os.chdir(cwd)
    print("Current directory -", os.getcwd())
    filetype = option.get()    
    root.update()
    
    if filetype == 'AM Accounting':
        try:
            file = ctk.filedialog.askopenfilename(title="Select a file", initialdir=cwd)
            excelRunner(file)
        except Exception as error:
            textbox.insert(ctk.END, f"ERROR: {error}\n")
                
    elif filetype == 'PM Accounting':
        try:
            textbox.insert(ctk.END, f"Starting PM Accounting Process!\n")
            PM_Run()
            textbox.insert(ctk.END, f"Process Complete!\n")
        except Exception as error:
            textbox.insert(ctk.END, f"ERROR: {error}\n")
        
    
    textbox.configure(state="disabled")
 
    
gui()
root.mainloop()